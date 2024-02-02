import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from pathlib import Path

def shape_to_excel_range(df_shape: tuple[int, int]) -> str:
    '''
    Given a tuple which represents the shape of the pandas data frame, this function
    converts it to an excel range represented in a string format.

    For example: (5, 5) -> "A1:E5"

    This is useful to convert a sheet into an excel table using openpyxl

    Parameters
    ----------
    df_shape: tuple representing the shape of the dataframe

    Returns
    -------
    a str object for excel
    '''
    
    if len(df_shape) != 2:
        raise ValueError("DataFrame shape must be a tuple of length 2 (rows, columns)")

    rows, cols = df_shape
    start_cell = 'A1'
    end_cell = get_column_letter(cols) + str(rows)

    return f"{start_cell}:{end_cell}"


def excel_writer(writer: pd.ExcelWriter, dfs: dict[str, pd.DataFrame], to_tables: dict[str, bool]):
    '''
    Write multiple dataframes to the given sheets using an excel writer object created using pd.ExcelWriter

    Parameters
    ----------

    writer: pd.ExcelWriter
        Excel writer object to write in a excel workbook

    dfs: dict[str, pd.DataFrame]
        A dictionary object where key is the sheet name and value is the dataframe

    to_tables: dict[str, bool]
        A dictionary object where key is the sheet name and value is a True or False 
        specifying whether or not the given sheet should be converted to an Excel Table.

    Returns
    -------
    None
    '''
    for sheet_name, df in dfs.items():
        convert_to_table = to_tables[sheet_name]
        print(f'\tWriting df.shape: {df.shape} to {sheet_name}: ', end='', flush=True)
        try:
            # Try to write the DataFrame to the specified sheet
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            if convert_to_table:
                tab = Table(displayName=sheet_name, ref=shape_to_excel_range(df.shape))
                # Convert the data to an Excel table
                writer.sheets[sheet_name].add_table(tab)
        except Exception as e:
            print(f"\tError writing to sheet {sheet_name}: {e}")
    
        print('Done')

def write_dataframe_to_excel(filename: Path | str, dfs: dict[str, pd.DataFrame], to_tables: dict[str, bool]=None):
    '''
    Creates a writer object and passes it to excel_writer function.
    Write multiple dataframes to the given sheets using an excel writer object created using pd.ExcelWriter

    Parameters
    ----------

    writer: pd.ExcelWriter
        Excel writer object to write in a excel workbook

    dfs: dict[str, pd.DataFrame]
        A dictionary object where key is the sheet name and value is the dataframe

    to_tables: dict[str, bool]
        A dictionary object where key is the sheet name and value is a True or False 
        specifying whether or not the given sheet should be converted to an Excel Table.

    Returns
    -------
    None
    '''
    filename = Path(filename).resolve()
    try:
        # Load the existing Excel file
        with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            print(f'Opening existing workbook: {filename.name}', end='\n')
            excel_writer(writer, dfs, to_tables)
    except FileNotFoundError:
        print(f'Creating a new workbook: {filename.name}', end='\n')
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            excel_writer(writer, dfs, to_tables)